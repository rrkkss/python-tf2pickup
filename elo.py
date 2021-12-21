import openpyxl

base_elo = 1600
nicknames_global = {}

red_team = []
red_team_tempElo  = []
red_team_tempEloBonus = []
red_score = 0
red_avg_elo = 0
red_win_score_factor = 0.0

blue_team = []
blue_team_tempElo = []
blue_team_tempEloBonus = []
blue_score = 0
blue_avg_elo = 0
blue_win_score_factor = 0.0

try:
    wb = openpyxl.load_workbook('elo.xlsx')
except:
    print('nt')

for i in wb.sheetnames:
    if i != 'elo':
        wb.create_sheet(title='elo')

def EloCreateNewName(player_nick):
    column = 1
    limit = 0
    countRows = 1

    if wb['elo'].cell(1,1).value is None:
        wb['elo'].cell(1,1).value = player_nick
        wb['elo'].cell(1,2).value = base_elo
        return 0

    while limit < 1:
        if wb['elo'].cell(countRows,1).value == player_nick:
            return 0
        if wb['elo'].cell(countRows,1).value is None:
            break
        countRows += 1

    wb['elo'].cell(countRows, column).value = player_nick
    wb['elo'].cell(countRows, column+1).value = base_elo

def EloCalculationPrepare():
    limit = 0
    red_avg_elo = 0.0
    blue_avg_elo = 0.0
    
    # RED
    for i in range(0,len(red_team),1):
        limit = 0
        rows_method = 1
        while limit < 1:
            if wb['elo'].cell(rows_method, 1).value == red_team[i]:
                red_avg_elo += wb['elo'].cell(rows_method, 2).value
                limit = 1
            rows_method += 1
    
    # BLUE
    for i in range(0,len(blue_team),1):
        limit = 0
        rows_method = 1
        while limit < 1:
            if wb['elo'].cell(rows_method, 1).value == blue_team[i]:
                blue_avg_elo += wb['elo'].cell(rows_method, 2).value
                limit = 1
            rows_method += 1

    if red_score > blue_score:
        red_win_score_factor = 1
        blue_win_score_factor = 0
        EloCalculation(red_win_score_factor, blue_win_score_factor, red_avg_elo/6, blue_avg_elo/6)
    
    elif red_score == blue_score:
        red_win_score_factor = 0.5
        blue_win_score_factor = 0.5
        EloCalculation(red_win_score_factor, blue_win_score_factor, red_avg_elo/6, blue_avg_elo/6)
        
    elif red_score < blue_score:
        red_win_score_factor = 0
        blue_win_score_factor = 1
        EloCalculation(red_win_score_factor, blue_win_score_factor, red_avg_elo/6, blue_avg_elo/6)

def EloCalculation(rF, bF, rAvg, bAvg):
    limit = 0
    red_team_tempElo.clear()
    blue_team_tempElo.clear()
    win_chance = 0.0

    if rF > bF: #vyhra cervenych + prohra modrych
        for i in range(0,len(red_team),1): #cervena
            limit = 0
            rows_method = 1
            while limit < 1:
                if wb['elo'].cell(rows_method, 1).value == red_team[i]:
                    eloC = wb['elo'].cell(rows_method, 2).value
                    limit = 1
                rows_method += 1

            win_chance = 1/(1+10**((bAvg - eloC)/400)) #oponent minus hracovo elo, hrac cerveny odecita se od modre
            eloN = eloC + 32*(1 - win_chance)
            red_team_tempElo.append(eloN)

        for i in range(0,len(blue_team),1): #modra
            limit = 0
            rows_method = 1
            while limit < 1:
                if wb['elo'].cell(rows_method, 1).value == blue_team[i]:
                    eloC = wb['elo'].cell(rows_method, 2).value
                    limit = 1
                rows_method += 1
        
            win_chance = 1/(1+10**((rAvg - eloC)/400)) #oponent minus hracovo elo, hrac modre odecita se od cervene
            eloN = eloC + 32*(0 - win_chance)
            blue_team_tempElo.append(eloN)
    
    if rF == bF: #remiza cervenych + remize modrych
        for i in range(0,len(red_team),1): #cervena
            limit = 0
            rows_method = 1
            while limit < 1:
                if wb['elo'].cell(rows_method, 1).value == red_team[i]:
                    eloC = wb['elo'].cell(rows_method, 2).value
                    limit = 1
                rows_method += 1

            win_chance = 1/(1+10**((bAvg - eloC)/400)) #oponent minus hracovo elo, hrac cerveny odecita se od modre
            eloN = eloC + 32*(0.5 - win_chance)
            red_team_tempElo.append(eloN)

        for i in range(0,len(blue_team),1): #modra
            limit = 0
            rows_method = 1
            while limit < 1:
                if wb['elo'].cell(rows_method, 1).value == blue_team[i]:
                    eloC = wb['elo'].cell(rows_method, 2).value
                    limit = 1
                rows_method += 1
        
            win_chance = 1/(1+10**((rAvg - eloC)/400)) #oponent minus hracovo elo, hrac modre odecita se od cervene
            eloN = eloC + 32*(0.5 - win_chance)
            blue_team_tempElo.append(eloN)

    if rF < bF: #remiza cervenych + remize modrych
        for i in range(0,len(red_team),1): #cervena
            limit = 0
            rows_method = 1
            while limit < 1:
                if wb['elo'].cell(rows_method, 1).value == red_team[i]:
                    eloC = wb['elo'].cell(rows_method, 2).value
                    limit = 1
                rows_method += 1

            win_chance = 1/(1+10**((bAvg - eloC)/400)) #oponent minus hracovo elo, hrac cerveny odecita se od modre
            eloN = eloC + 32*(0 - win_chance)
            red_team_tempElo.append(eloN)

        for i in range(0,len(blue_team),1): #modra
            limit = 0
            rows_method = 1
            while limit < 1:
                if wb['elo'].cell(rows_method, 1).value == blue_team[i]:
                    eloC = wb['elo'].cell(rows_method, 2).value
                    limit = 1
                rows_method += 1
        
            win_chance = 1/(1+10**((rAvg - eloC)/400)) #oponent minus hracovo elo, hrac modre odecita se od cervene
            eloN = eloC + 32*(1 - win_chance)
            blue_team_tempElo.append(eloN)

def EloWrite():
    for i in range(0,len(red_team),1): #cervena
        limit = 0
        rows_method = 1
        while limit < 1:
            if wb['elo'].cell(rows_method, 1).value == red_team[i]:
                try:
                    wb['elo'].cell(rows_method, 2).value = red_team_tempElo[i] + red_team_tempEloBonus[i]
                except Exception as e:
                    print('fungovat by to melo, ale obcas to hodi index out of range')
                    print('chyba:',e)
                limit = 1
            rows_method += 1

    for i in range(0,len(blue_team),1): #modra
        limit = 0
        rows_method = 1
        while limit < 1:
            if wb['elo'].cell(rows_method, 1).value == blue_team[i]:
                try:
                    wb['elo'].cell(rows_method, 2).value = blue_team_tempElo[i] + blue_team_tempEloBonus[i]
                except Exception as e:
                    print('fungovat by to melo, ale obcas to hodi index out of range')
                    print('chyba:',e)
                limit = 1
            rows_method += 1

def EloConvertIDtoNicks():
    wb = openpyxl.load_workbook('elo.xlsx')

    limit = 0
    rows = 1
    count = 0
    while limit < 1:
        count = 0
        if str(wb['elo'].cell(rows, 1).value)[0] == "[":
            for k,v in nicknames_global.items():
                count += 1
                if wb['elo'].cell(rows, 1).value == k:
                    wb['elo'].cell(rows, 1).value = v

        elif limit < -10:
            print('nt')
        
        else:
            limit += 1

        rows += 1

    try:
        wb.save('elo.xlsx')
        print('uspesne ulozeno a prepsany ID na nicky!')
    except Exception as e:
        print(e)

def EloMain(js):
    red_team.clear()
    red_team_tempElo.clear()
    red_team_tempEloBonus.clear()
    blue_team.clear()
    red_team_tempElo.clear()
    blue_team_tempEloBonus.clear()

    player_cycle = 0
    name_cycle = 0

    ns = js['names']
    
    game_lenght = int(js['length'])

    nicknames = dict(ns)
    nicknames_global.update(nicknames)

    for x in js["teams"].items():
        if x[0] == 'Red':
            red_score = x[1]["score"]
        elif x[0] == 'Blue':
            blue_score = x[1]["score"]

    for o in js['players'].items(): #o je kazdy ID hrace
        for i in nicknames.keys():
            if o[0] == i:
                #player_nick = nicknames[o[0]] #priradi se aktualni nick
                player_nick = o[0] #nepriradi se, zustane na ID
                #print('ID hrace:',o[0],'nick:',player_nick)
                
            elif o[0] != i:
                name_cycle += 1
                continue

            else:
                print('~~~~chyba pri kontrole jmen, nesedi id nebo neco')
                print('skipuje se\n')
                name_cycle += 1
                continue
        
        EloCreateNewName(player_nick) #kdyz neni hrac zarazen, zde se vytvori se defaultnim elem 1600

        player_info = o[1].items()
        player_info = dict(player_info)
        player_class = list(str(player_info["class_stats"][0]).split("'"))[3]
        player_kpd = player_info["kpd"] #kills per death    
        player_kapd = player_info["kapd"] #kills+assits per death
        
        player_damage = player_info["dmg"]
        player_dt = player_info["dt"] #dmg taken
        player_dpm = player_info["dapm"] #dmg per minute
        player_captures = player_info["cpc"]

        player_heals = player_info["heal"]

        log_team = o[1].items()
        log_team = dict(log_team)
        log_team = log_team["team"]
        
        if log_team == 'Blue':
            blue_team.append(player_nick)

            if player_class == 'scout':
                extra_elo = (float(player_kapd) - 1.7)*5 + (float(player_dpm) - 240)/20 + float(player_captures)/5
                blue_team_tempEloBonus.append(extra_elo)
            
            elif player_class == 'soldier':
                extra_elo = (float(player_kpd) - 1)*5 + (float(player_dpm) - 280)/20 + (float(player_damage) - float(player_dt))/500
                blue_team_tempEloBonus.append(extra_elo)

            elif player_class == 'demoman':
                extra_elo = (float(player_kpd) - 1.5)*7 + (float(player_dpm) - 330)/20 + (float(player_kapd) - 1.5)*2
                blue_team_tempEloBonus.append(extra_elo)

            elif  player_class == 'medic':
                extra_elo = (float(player_kapd) - 1.5)*5 + (float(player_heals))**(1/5) + (float(player_heals / (game_lenght / 60.0)) - 800)/25
                blue_team_tempEloBonus.append(extra_elo)

        elif log_team == 'Red':
            red_team.append(player_nick)

            if player_class == 'scout':
                extra_elo = (float(player_kapd) - 1.7)*5 + (float(player_dpm) - 240)/20 + float(player_captures)/5
                red_team_tempEloBonus.append(extra_elo)
            
            elif player_class == 'soldier':
                extra_elo = (float(player_kpd) - 1)*5 + (float(player_dpm) - 280)/20 + (float(player_damage) - float(player_dt))/500
                red_team_tempEloBonus.append(extra_elo)

            elif player_class == 'demoman':
                extra_elo = (float(player_kpd) - 1.5)*7 + (float(player_dpm) - 330)/20 + (float(player_kapd) - 1.5)*2
                red_team_tempEloBonus.append(extra_elo)

            elif  player_class == 'medic':
                extra_elo = (float(player_kapd) - 1.5)*5 + (float(player_heals))**(1/5) + (float(player_heals / (game_lenght / 60.0)) - 800)/25
                red_team_tempEloBonus.append(extra_elo)

        player_cycle += 1

    #print(red_team)
    #print(blue_team)
    #tady je check na elo, protoze konec logu mapky
    EloCalculationPrepare()
    EloWrite()
    
    #log_cycle += 1
    #if log_cycle == 3:
        #break