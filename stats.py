import openpyxl
import xlsxwriter
import time

try:
    wb = openpyxl.load_workbook('stats.xlsx')
except:
    print("file doesnt exist, creating new one 'stats.xlsx'")
    try:
        workbook = xlsxwriter.Workbook('stats.xlsx')
        worksheet = workbook.add_worksheet('title')
        workbook.close()
        wb = openpyxl.load_workbook('stats.xlsx')
    except:
        print('closing in 5 seconds')
        time.sleep(5)
        quit()

column = 0

def StatsHeader(active_sheet):
    row = 2
    for i in range(0,4,1):
        if i == 0:
            column = 2
            player_class_header = 'scout'
        if i == 1:
            column = 8
            player_class_header = 'soldier'
        elif i == 2:
            column = 15
            player_class_header = 'demoman'
        elif i == 3:
            column = 22
            player_class_header = 'medic'
        wb[active_sheet].cell(row, column).value = 'class:'
        wb[active_sheet].cell(row, column+1).value = player_class_header
        wb[active_sheet].cell(row, column+2).value = 'games:'
        wb[active_sheet].cell(row, column+3).value = '0'

    for i in range(0,4,1):
        if i == 0: #scout, column 2
            data_header = ['kills', 'assists', 'deaths', 'DMG', 'DPM']
            StatsHeaderWrite(active_sheet, 2, data_header)

        if i == 1: #soldier, column 8
            data_header = ['kills', 'assists', 'deaths', 'DMG', 'DPM', 'airshots']
            StatsHeaderWrite(active_sheet, 8, data_header)

        elif i == 2: #demo, column 15
            StatsHeaderWrite(active_sheet, 15, data_header)

        elif i == 3: #medic, column 22
            data_header = ['kills', 'assists', 'deaths', 'DMG', 'DPM', 'ubers', 'heals']
            StatsHeaderWrite(active_sheet, 22, data_header)

def StatsHeaderWrite(active_sheet, clmn, header):
    row = 3
    items = 0

    for i in header:
        items += 1
    for i in range(0,items,1):
        wb[active_sheet].cell(row, clmn).value = header[i]
        clmn += 1

def StatsClassWrite(name, pclass, game_count, stats):
    row = game_count + 5

    if pclass == 'scout':
        column = 2
        for i in range(0,5,1):
            wb[name].cell(row, column+i).value = stats[i]
    elif pclass == 'soldier':
        column = 8
        for i in range(0,6,1):
            wb[name].cell(row, column+i).value = stats[i]
    elif pclass == 'demoman':
        column = 15
        for i in range(0,6,1):
            wb[name].cell(row, column+i).value = stats[i]
    elif pclass == 'medic':
        column = 22
        for i in range(0,7,1):
            wb[name].cell(row, column+i).value = stats[i]

def StatsBase(js, count, results):
    if count < results:
        ns = js['names']
        nicknames = dict(ns)
        #print(nicknames)

        player_cycle = 0
        name_cycle = 0
        column = 0

        for o in js['players'].items(): #o is ID of a player
            for i in nicknames.keys():
                if o[0] == i:
                    player_nick = nicknames[o[0]]
                    #print('ID hrace:',o[0],'nick:',player_nick)
                    
                elif o[0] != i:
                    name_cycle += 1
                    continue

                else:
                    print('~~~~error with checking names, ID doesnt seem to match')
                    print('skipping\n')
                    name_cycle += 1
                    continue
            
            player_info = o[1].items()
            player_info = dict(player_info)
                    
            player_class = list(str(player_info["class_stats"][0]).split("'"))[3] #this is a bit wonky
                #offclasses do get counted in the main class, but this was done due to offclassing being minimal          
            player_kills = player_info["kills"]
            player_deaths = player_info["deaths"]
            player_assists = player_info["assists"]
            
            player_damage = player_info["dmg"]
            player_dpm = player_info["dapm"]

            # scout; soldier; demo; medic
            if player_class == 'scout':
                player_heals = 0
                player_ubers = 0
                player_airshots = 0
                stats = [player_kills, player_assists, player_deaths, player_damage, player_dpm]

            elif player_class == 'soldier' or player_class == 'demoman':
                player_heals = 0
                player_ubers = 0
                player_airshots = player_info["as"]
                stats = [player_kills, player_assists, player_deaths, player_damage, player_dpm, player_airshots]

            elif player_class == 'medic':
                player_heals = player_info["heal"] #other classes have this as how many heals they obtained
                player_ubers = player_info["ubers"]
                player_airshots = 0
                stats = [player_kills, player_assists, player_deaths, player_damage, player_dpm, player_ubers, player_heals]

            nicks = []
            instances_ofNick = 0
            for i in wb.sheetnames:
                nicks.append(i)

            instances_ofNick = nicks.count(player_nick)
            if instances_ofNick == 0:
                try:
                    wb.create_sheet(title=player_nick)
                    StatsHeader(player_nick) #last added sheet is always active
                except Exception as e:
                    print(e)
            elif instances_ofNick > 1:
                print('player is categorized 2 times')

            for i in range(len(wb.sheetnames)):
                if player_nick == wb.sheetnames[i]:
                    
                    if player_class == 'scout':
                        games_count = int(wb[player_nick].cell(2, 5).value)
                        wb[player_nick].cell(2, 5).value = games_count + 1

                    elif player_class == 'soldier':
                        games_count = int(wb[player_nick].cell(2, 11).value)
                        wb[player_nick].cell(2, 11).value = games_count + 1

                    elif player_class == 'demoman':
                        games_count = int(wb[player_nick].cell(2, 18).value)
                        wb[player_nick].cell(2, 18).value = games_count + 1

                    elif player_class == 'medic':
                        games_count = int(wb[player_nick].cell(2, 25).value)
                        wb[player_nick].cell(2, 25).value = games_count + 1

                    try:
                        StatsClassWrite(player_nick, player_class, games_count, stats)
                    except Exception as e:
                        print('~~~~ player info couldnt be written, skipping: ',e)
                        continue
                
            player_cycle += 1

    elif count == results:
        try:
            wb.save('stats.xlsx')
            print("succesfully saved into 'stats.xlsx'")

        except Exception as e:
            print(e)

    else:
        print('something unexpected happened')