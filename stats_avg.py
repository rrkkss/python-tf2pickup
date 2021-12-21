import openpyxl
from openpyxl.styles import PatternFill

try:
    wb = openpyxl.load_workbook('stats.xlsx')
except:
    print('lol')

columns = [2, 8, 15, 22] #scout, soldier, demo, medic

def StatsReadSheet(column, sheet):
    
    while wb[sheet].cell(5, column).value != None:
        total = 0.0
        row = 5
        while wb[sheet].cell(row, column).value != None:
            total += wb[sheet].cell(row, column).value
            row += 1

        try:
            if column >= 2 and column <= 6:
                wb[sheet].cell(4, column).value = total / int(wb[sheet].cell(2,5).value)
                wb[sheet].cell(4, column).fill = PatternFill(start_color="0ff1ce", fill_type="solid")
                
            elif column >= 8 and column <= 13:
                
                wb[sheet].cell(4, column).value = total / int(wb[sheet].cell(2,11).value)
                wb[sheet].cell(4, column).fill = PatternFill(start_color="B88B69", fill_type="solid")
            

            elif column >= 15 and column <= 20:
                
                wb[sheet].cell(4, column).value = total / int(wb[sheet].cell(2,18).value)
                wb[sheet].cell(4, column).fill = PatternFill(start_color="F66B7B", fill_type="solid")
                

            elif column >= 22 and column <= 28:
                wb[sheet].cell(4, column).value = total / int(wb[sheet].cell(2,25).value)
                wb[sheet].cell(4, column).fill = PatternFill(start_color="AAAAAA", fill_type="solid")

        except (Exception, ZeroDivisionError) as e:
            pass #this just spams the console, many players havent played all the classes    

        column += 1

def StatsAVGcreate():
    for sheet in wb.sheetnames:
        if sheet == "title":
            continue
        for n in columns:
            StatsReadSheet(n,sheet)

    try:
        wb.save('stats.xlsx')
        print("succesfully saved into 'stats.xlsx', averages are done")
    except Exception as e:
        print(e)

StatsAVGcreate()
