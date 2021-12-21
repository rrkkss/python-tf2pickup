import openpyxl
from openpyxl import load_workbook

wb = openpyxl.load_workbook('elo.xlsx')
newElo_list = []

elo_list = []
name_list = {}
count = 0

def createList():

    limit = 0
    rows = 1
    while limit < 1:
        if wb['elo'].cell(rows, 2).value != None:
            newValue = wb['elo'].cell(rows, 2).value
            elo_list.append(newValue)

            name_list[wb['elo'].cell(rows, 2).value] = wb['elo'].cell(rows, 1).value
            #pouziva elo jako klic, jmeno je value

        elif limit < -10:
            print('nt')
        
        else:
            limit += 1

        rows += 1

def bubblesort(list):

# Swap the elements to arrange in order
    for iter_num in range(len(list)-1,0,-1):
        for idx in range(iter_num):
            if list[idx]<list[idx+1]: #predtim >
                temp = list[idx]
                list[idx] = list[idx+1]
                list[idx+1] = temp
    
    return list

def appendNamesToElo(list):
    limit = 0
    rows = 1
    while limit < 1:
        for k,v in list.items():
            if wb['elo'].cell(rows, 2).value == k:
                wb['elo'].cell(rows, 1).value = v
            
            elif wb['elo'].cell(rows, 2).value == None:
                limit = 1

        rows += 1

    try:
        wb.save('elo.xlsx')
        print('uspesne ulozeno a prepsany ID na nicky!')
    except Exception as e:
        print(e)

#createList()
#newElo_list = bubblesort(elo_list)

#for i in newElo_list:
#    count += 1
#    wb['elo'].cell(count, 2).value = i

#try:
#    wb.save('elo.xlsx')
#    print('uspesne ulozeno!')
#except Exception as e:
#    print(e)

#appendNamesToElo(name_list)