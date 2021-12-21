import requests, json, time, os
#import urllib.request
#import openpyxl
from openpyxl import load_workbook
import xlsxwriter

import elo
import elo_sort
import maps
import stats
import stats_avg

log_list = []
log_cycle = 0

entryLoop = False
entryXlsxSave = False

def SaveWorkbook():
    try:
        workbook.close()
    except Exception as e:
        print(e)

def CheckForFile(name):
    if os.path.exists(name) == False:
        return False
    elif os.path.exists(name) == True:
        print('File exists, not creating a new one')
        return True

def CheckEnd():
    if count == row-1:
        return True
    return False

def CalcTimeSubMethod(input):
    result = int(input) * user_input_sleep_time
    resultMin = int(result / 60)
    resultSec = round(result % 60, 2)
    
    result = str(resultMin) + " min " + str(resultSec) + " sec "
    return result

def CalcTime(current, total):
    tbd = CalcTimeSubMethod(current)
    ttl = CalcTimeSubMethod(total)

    return " | " + tbd + "/ " + ttl

while entryLoop == False:
    print('\n it is advised to delete both elo.xlsx and stats.xlsx before using this program,') 
    print('as there is currently no last game state (all games will be appended to the existing files)\n')

    print("'elo' for player elos, 'stats' for player class statistics and 'maps' for number of maps played")       
    user_input_jsonType = input("Enter 'elo', 'stats' or 'maps' -> ")
    if user_input_jsonType == "elo" and CheckForFile(user_input_jsonType) == False:
        workbook = xlsxwriter.Workbook('elo.xlsx')
        worksheet = workbook.add_worksheet('elo')
        SaveWorkbook()
        entryLoop = True

    elif user_input_jsonType == "stats" and CheckForFile(user_input_jsonType) == False:
        workbook = xlsxwriter.Workbook('stats.xlsx')
        worksheet = workbook.add_worksheet('title')
        SaveWorkbook()
        entryLoop = True

    elif user_input_jsonType == "maps":
        entryLoop = True

    elif CheckForFile(user_input_jsonType) == True:
        entryLoop = True

    else:
        print("bad entry")

user_input_site = input("Enter logs.tf site disclosure, eg for czech pickup it is 'tf2pickup.cz', for german 'tf2pickup.de' -> ")
user_input_sleep_time = input("Enter number of seconds to wait in between logs, do not use zero as it gets caught by ddos protection. good value is 0.3 -> ")
user_input_sleep_time = float(user_input_sleep_time)

url = 'http://logs.tf/api/v1/log?title=' + user_input_site
print('parsing from:',url)

try:
    r = requests.get(url)
    j = r.json()
except Exception:
    print('json couldnt be parsed, closing in 5 seconds')
    time.sleep(5)
    quit()

res = list(j.values())
row = res[1]+1
if user_input_site == 'tf2pickup.cz':
    row = res[1]-23 # reason is in the comment below, it was 24 of them
print('results:',row-1,'\n')

for i in j:
    if i == 'logs':
        for k in j[i]:
            idn = list(k.values())
            if user_input_site == 'tf2pickup.cz':
                if int(idn[0]) < 2865412: #first actual czech pug, before that dmixes were under the same name
                    break
            log_list.append(idn[0])

print(log_list)

count = 0
for i in log_list:
    count += 1
    time.sleep(user_input_sleep_time)
    url = 'https://logs.tf/json/'
    url = url + str(log_list[log_cycle])
    print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
    
    try:
        r = requests.get(url)
        js = r.json()
    except Exception as e:
        print('~~~bad response or mistake in parsing, skipping, reason:', e)
        continue
        
    del js['chat'] #fuck rgl
        
    if js['success'] == True: #just in case
        print(url,'OK')
        print(count, '/',row-1)#, CalcTime(count, row-1))
    else:
        print('~~~~ log is broken (broken on logs.tf page) \n')
        continue

    if user_input_jsonType == "elo":
        elo.EloMain(js)
        if CheckEnd() == True:
            elo.EloConvertIDtoNicks()
            inp = input("would you like to sort the list? 'Y' / 'n' -> ")
            if inp == "Y":
                elo_sort.createList()
                elo_sort.newElo_list = elo_sort.bubblesort(elo_sort.elo_list)

                for i in elo_sort.newElo_list:
                    count += 1
                    elo_sort.wb['elo'].cell(count, 2).value = i

                    try:
                        elo_sort.wb.save('elo.xlsx')
                        print('uspesne ulozeno!')
                    except Exception as e:
                        print(e)

                elo_sort.appendNamesToElo(elo_sort.name_list)
            
            else:
                quit()

    elif user_input_jsonType == "stats":
        stats.StatsBase(js, count, row-1)
        if CheckEnd() == True:
            inp = input("would you like to run avgs? 'Y' / 'n' CURRENTLY BROKEN, RUN AVG AFTER THIS PROGRAM ENDS ")
            if inp == "Y":
                stats_avg.StatsAVGcreate()
            else:
                quit()

    elif user_input_jsonType == "maps":
        maps.MapsCount(js)
        if CheckEnd() == True:
            maps.MapsPrint()

    log_cycle += 1